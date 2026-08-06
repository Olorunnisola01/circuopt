"""Parsers for the two open datasets CircuOpt is built on.

Both are redistributable without restriction:

* UK Government GHG Conversion Factors for Company Reporting 2025
  (DESNZ/DEFRA), published under the Open Government Licence v3.0.
  File: data/ghg-conversion-factors-2025-flat-format.xlsx
  Gives, per material, a *primary material production* factor and a
  *closed-loop source* factor in kg CO2e per tonne, plus a factor per
  end-of-life route, plus UK grid electricity generation and T&D factors.

* USGS Mineral Commodity Summaries 2025 salient-statistics data releases,
  a work of the U.S. Government and therefore public domain.
  Files: data/usgs/mcs2025-*_salient.csv
  Gives annual average commodity prices used for the life-cycle costing.

Nothing here invents a number.  Every value the optimiser consumes is
either read out of one of these two files or declared in factors.py with
an explicit source and an explicit "assumption" label.
"""

from __future__ import annotations

import csv
import json
import os
import re
from dataclasses import dataclass, field

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(os.path.dirname(HERE), "data")

DEFRA_XLSX = os.path.join(DATA, "ghg-conversion-factors-2025-flat-format.xlsx")
USGS_DIR = os.path.join(DATA, "usgs")


# --------------------------------------------------------------------------
# DEFRA
# --------------------------------------------------------------------------

@dataclass
class DefraFactors:
    """kg CO2e per kilogram, keyed by DEFRA 'Level 3' material name."""

    primary: dict = field(default_factory=dict)      # virgin production
    closed_loop: dict = field(default_factory=dict)  # recycled-content input
    disposal: dict = field(default_factory=dict)     # {material: {route: kgCO2e/kg}}
    electricity_generation: float = 0.0              # kg CO2e per kWh
    electricity_td: float = 0.0                      # kg CO2e per kWh, T&D losses
    source: str = ""

    @property
    def electricity(self) -> float:
        return self.electricity_generation + self.electricity_td


def load_defra(path: str = DEFRA_XLSX) -> DefraFactors:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Factors by Category"]
    rows = list(ws.iter_rows(min_row=7, values_only=True))

    f = DefraFactors(source=os.path.basename(path))
    for r in rows:
        _id, _scope, l1, _l2, l3, _l4, col, uom, ghg, val = r[:10]
        if val is None or ghg != "kg CO2e":
            continue

        if l1 == "Material use" and uom == "tonnes":
            per_kg = val / 1000.0
            if col == "Primary material production":
                f.primary[l3] = per_kg
            elif col == "Closed-loop source":
                f.closed_loop[l3] = per_kg

        elif l1 == "Waste disposal" and uom == "tonnes":
            f.disposal.setdefault(l3, {})[col] = val / 1000.0

        elif l1 == "UK electricity" and l3 == "Electricity: UK" and uom == "kWh":
            f.electricity_generation = val
        elif l1 == "Transmission and distribution" and l3 == "Electricity: UK" \
                and uom == "kWh":
            f.electricity_td = val

    if not f.primary or f.electricity_generation == 0.0:
        raise RuntimeError("DEFRA workbook parsed but expected rows are missing")
    return f


# --------------------------------------------------------------------------
# USGS
# --------------------------------------------------------------------------

LB = 0.45359237
TROY_OZ = 0.0311034768

#: USGS encodes the price unit as a suffix on the column name.
_UNIT_SUFFIX = {
    "ctslb": (0.01 / LB, "cents/lb"),
    "dlb": (1.0 / LB, "USD/lb"),
    "dt": (1.0 / 1000.0, "USD/t"),
    "dmt": (1.0 / 1000.0, "USD/t"),
    "dtoz": (1.0 / TROY_OZ, "USD/troy oz"),
    "dkg": (1.0, "USD/kg"),
}

#: Which price column to prefer when a commodity reports several.  LME and
#: COMEX quotes are excluded in favour of the US-market series so that all
#: commodities are on a comparable basis.
_PREFERRED = ("Price_US_ctslb", "Price_ctslb", "Price_Num_1_heavy_dt",
              "Price_dt", "Price_Bullion_dtoz", "Price_NY_ctslb")


def _num(tok: str):
    tok = (tok or "").strip().replace("$", "").replace(",", "")
    if not tok or tok in {"NA", "W", "--", "XX", "E", "e"}:
        return None
    try:
        return float(tok)
    except ValueError:
        return None


def load_usgs_prices(directory: str = USGS_DIR) -> dict:
    """Return ``{commodity: {...}}`` with the latest annual price in USD/kg.

    Each MCS 2025 salient-statistics file is a tidy table with one row per
    year and one or more ``Price_*`` columns whose unit is encoded in the
    column-name suffix (``ctslb`` = cents per pound, ``dt`` = dollars per
    metric ton, ``dtoz`` = dollars per troy ounce, ...).
    """
    out = {}
    if not os.path.isdir(directory):
        return out

    for name in sorted(os.listdir(directory)):
        if not name.endswith("_salient.csv"):
            continue
        path = os.path.join(directory, name)
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            rows = list(csv.DictReader(fh))
        if not rows or "Commodity" not in rows[0]:
            continue

        price_cols = [c for c in rows[0] if c and c.startswith("Price_")]
        col = next((c for c in _PREFERRED if c in price_cols), None)
        if col is None and price_cols:
            col = price_cols[0]
        if col is None:
            continue

        suffix = col.rsplit("_", 1)[-1].lower()
        if suffix not in _UNIT_SUFFIX:
            continue
        factor, unit = _UNIT_SUFFIX[suffix]

        series = [(r.get("Year", ""), _num(r.get(col))) for r in rows]
        series = [(y, v) for y, v in series if v is not None]
        if not series:
            continue
        year, value = series[-1]

        out[rows[0]["Commodity"].strip()] = {
            "usd_per_kg": value * factor,
            "reported_value": value,
            "reported_unit": unit,
            "column": col,
            "year": year,
            "history": {y: v * factor for y, v in series},
            "file": name,
        }
    return out


def dump_cache(out_path: str) -> dict:
    d = load_defra()
    payload = {
        "defra": {
            "primary_kgco2e_per_kg": d.primary,
            "closed_loop_kgco2e_per_kg": d.closed_loop,
            "disposal_kgco2e_per_kg": d.disposal,
            "electricity_kgco2e_per_kwh": d.electricity,
            "electricity_generation": d.electricity_generation,
            "electricity_td": d.electricity_td,
            "source": d.source,
        },
        "usgs_prices": load_usgs_prices(),
    }
    with open(out_path, "w") as fh:
        json.dump(payload, fh, indent=1, sort_keys=True)
    return payload


if __name__ == "__main__":
    p = dump_cache(os.path.join(DATA, "factors_cache.json"))
    d = p["defra"]
    print("DEFRA materials with a primary factor :", len(d["primary_kgco2e_per_kg"]))
    print("DEFRA materials with a closed-loop one:", len(d["closed_loop_kgco2e_per_kg"]))
    print("UK grid factor (gen + T&D)            :",
          round(d["electricity_kgco2e_per_kwh"], 5), "kg CO2e/kWh")
    print("USGS commodities priced               :", len(p["usgs_prices"]))
    for k, v in sorted(p["usgs_prices"].items()):
        print(f"   {k:22s} {v['usd_per_kg']:10.3f} USD/kg   ({v['reported_value']} "
              f"{v['reported_unit']}, {v['year']})")


# --------------------------------------------------------------------------
# Published bill of materials (GARO LS4 LCA, 2021)
# --------------------------------------------------------------------------

GARO_JSON = os.path.join(DATA, "garo_ls4_bom.json")


def load_garo_bom(path: str = GARO_JSON) -> dict:
    """The cited material distribution for the case product.

    Returns the parsed record plus a derived ``masses_kg`` mapping from
    material family to kilograms in one product, packaging excluded and
    renormalised onto the manufacturer's official 24.5 kg product mass.

    The report gives mass by material, not by component, so the split of each
    material across components is declared in the JSON and is this project's
    assumption; the per-family totals it must reproduce are not.
    """
    with open(path) as fh:
        rec = json.load(fh)

    dist = rec["material_distribution_pct_of_29_7kg"]
    packaging = set(rec["packaging_categories"])
    product_pct = {k: v for k, v in dist.items() if k not in packaging}
    total_pct = sum(product_pct.values())
    product_mass = rec["product"]["official_product_mass_kg"]

    # Map the report's material names onto the model's families.  Rubber is
    # folded into plastics; see the note in the JSON.
    family_of = {
        "Aluminium": "aluminium",
        "Electronic components": "electronics",
        "Copper/Plastic": "cable",
        "Steel": "steel",
        "Plastics": "plastic",
        "Rubber": "plastic",
    }
    masses = {}
    for name, pct in product_pct.items():
        fam = family_of[name]
        masses[fam] = masses.get(fam, 0.0) + product_mass * pct / total_pct

    rec["masses_kg"] = masses
    rec["product_mass_kg"] = product_mass
    rec["excluded_packaging_pct"] = round(
        sum(dist[k] for k in packaging), 3)
    return rec


def expand_bom(rec: dict | None = None) -> list:
    """Expand the cited family masses into component rows.

    Each family total is split by the fractions declared in
    ``allocation_to_components``; the fractions in each family sum to 1, so
    the component masses reproduce the published distribution exactly.
    Returns ``[(component_name, mass_kg, family), ...]``.
    """
    rec = rec or load_garo_bom()
    alloc = rec["allocation_to_components"]
    override = rec.get("component_family_override", {})
    out = []
    for fam, mass in rec["masses_kg"].items():
        shares = {k: v for k, v in alloc[fam].items() if not k.startswith("_")}
        s = sum(shares.values())
        if abs(s - 1.0) > 1e-6:
            raise ValueError(f"allocation for {fam} sums to {s}, not 1.0")
        for name, share in shares.items():
            out.append((name, mass * share, override.get(name, fam)))
    return out


# --------------------------------------------------------------------------
# World Bank "Pink Sheet" commodity prices -- a second, independent price
# source, used to test whether cost conclusions depend on which one is used.
# --------------------------------------------------------------------------

WORLDBANK_XLSX = os.path.join(DATA, "worldbank_cmo_monthly.xlsx")

#: World Bank column -> the model's commodity name, and the unit conversion
#: to USD/kg.  Both quoted commodities are in $/metric tonne.
_WB_COLUMNS = {
    "Aluminum": ("Aluminum", 1.0 / 1000.0),
    "Copper": ("Copper", 1.0 / 1000.0),
}


def load_worldbank_prices(path: str = WORLDBANK_XLSX) -> dict:
    """Return ``{commodity: {...}}`` from the World Bank Pink Sheet.

    A global reference-price series (the World Bank's own monthly
    commodity-market report), independent of the USGS Mineral Commodity
    Summaries used elsewhere in this project.  Steel/iron does not have a
    directly comparable entry in this source (it publishes iron ORE, not
    scrap, a different commodity), so the cross-check this dataset supports
    covers aluminium and copper only -- stated here rather than papered over
    with an approximate substitute.
    """
    if not os.path.exists(path):
        return {}
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Monthly Prices"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[4]

    out = {}
    for col_idx, cell in enumerate(header):
        if not cell:
            continue
        name = str(cell).strip()
        if name not in _WB_COLUMNS:
            continue
        commodity, factor = _WB_COLUMNS[name]
        series = []
        for r in rows[6:]:
            date, val = r[0], r[col_idx] if col_idx < len(r) else None
            if not date or not isinstance(val, (int, float)):
                continue
            series.append((str(date), float(val) * factor))
        if not series:
            continue
        last_date, last_val = series[-1]
        # a trailing 12-month average, so a single volatile month does not
        # dominate the comparison against USGS's annual figure
        last12 = [v for _, v in series[-12:]]
        out[commodity] = {
            "usd_per_kg_latest_month": last_val,
            "usd_per_kg_trailing_12m_mean": sum(last12) / len(last12),
            "latest_month": last_date,
            "n_months": len(series),
            "source": "World Bank Commodity Markets (Pink Sheet)",
        }
    return out
